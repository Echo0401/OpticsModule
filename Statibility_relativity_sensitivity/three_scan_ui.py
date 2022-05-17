# -*- coding: utf-8 -*-

import multiprocessing
import os
import sys
import threading

from PySide2 import QtWidgets, QtCore
from openpyxl import load_workbook
import numpy as np

from UI.loading import Loading
from UI.basis_dialog import Tip, Alert

from PyCRC.CRC32 import CRC32
from three_scan_function import ScanEDTest



crc = CRC32()


datapoints_report_dict = {'stability': {'fam':[-100, -100, [-100]], 'rox':[-100, -100, [-100]], 'hex':[-100, -100, [-100]], 'cy5':[-100, -100, [-100]]},
                          'relativity': {'fam':[-100, [-100]], 'rox':[-100, [-100]], 'hex':[-100, [-100]], 'cy5':[-100, [-100]]},
                          'sensitivity': {'fam':[-100, [-100]], 'rox':[-100, [-100]], 'hex':[-100, [-100]], 'cy5':[-100, [-100]]}}

def get_project_path():
    """
    获取项目的根目录
    :return: 根目录
    """
    # 判断调试模式
    debug_vars = dict((a, b) for a, b in os.environ.items() if a.find('IPYTHONENABLE') >= 0)
    # 根据不同场景获取根目录
    if len(debug_vars) > 0:
        """当前为debug运行时"""
        project_path = sys.path[2]
    elif getattr(sys, 'frozen', False):
        """当前为exe运行时"""
        project_path = os.getcwd()
    else:
        """正常执行"""
        project_path = sys.path[1]
    # 替换斜杠
    project_path = project_path.replace("\\", "/")
    return project_path


ProjectPath = get_project_path()

def get_file_absolute_path(fileName):
    """
    根据文件名获取资源文件路径
    """
    return ProjectPath + fileName

datapoint_template = get_file_absolute_path('/data/datapoint_template.xlsx')

class MainWindow(QtWidgets.QMainWindow):
    operation_signal = QtCore.Signal(bool, list)
    connect_signal = QtCore.Signal(bool)

    def __init__(self):
        super(MainWindow, self).__init__()
        version = "V1.0.1"
        self.scandedtest = ScanEDTest()
        self.communicate_status = False    # 确认串口的连接状态
        self.module_sn = ''              # 模块SN
        box_main = QtWidgets.QVBoxLayout()
        central_widget = QtWidgets.QWidget()
        central_widget.setLayout(box_main)
        self.setCentralWidget(central_widget)

        self.setWindowTitle(f"Scan-E1D1-E2D2-{version}")
        self.sn_module_edit = QtWidgets.QLineEdit()
        # self.module_sn = self.sn_module_edit.text()
        self.btn_search_com = QtWidgets.QPushButton("搜索模块")
        self.btn_search_com.clicked.connect(self.try_connect_module_com)
        self.label_connect_condition = QtWidgets.QLabel(f"<div>模 块 状 态：{self.module_sn} 未 连 接 <b><font size=5 color=red>o</font></b></div>")
        self.group_stability = QtWidgets.QGroupBox("稳定性")
        self.group_stability.setFixedWidth(300)
        self.group_relativity = QtWidgets.QGroupBox("截止率")
        self.group_relativity.setFixedWidth(300)
        self.group_sensitivity = QtWidgets.QGroupBox("灵敏度")
        self.group_sensitivity.setFixedWidth(300)
        self.btn_export_report = QtWidgets.QPushButton("导出报告")
        self.btn_export_report.clicked.connect(self.export_report)
        self.connect_signal.connect(self.connect_status)

        box_group = QtWidgets.QHBoxLayout()
        button_group = QtWidgets.QVBoxLayout()
        connect_group = QtWidgets.QVBoxLayout()
        connect_group.addWidget(self.label_connect_condition)
        edit_com_group = QtWidgets.QHBoxLayout()
        edit_com_group.addWidget(QtWidgets.QLabel("请输入模块SN号：  "))
        edit_com_group.addWidget(self.sn_module_edit)
        edit_com_group.addWidget(self.btn_search_com)
        box_group.addWidget(self.group_stability)
        box_group.addWidget(self.group_relativity)
        box_group.addWidget(self.group_sensitivity)
        button_group.addWidget(self.btn_export_report)
        box_main.addLayout(connect_group)
        box_main.addLayout(edit_com_group)
        box_main.addLayout(box_group)
        box_main.addLayout(button_group)

        # ------------稳定性布局--------------
        v_stability = QtWidgets.QVBoxLayout()

        f_stability = QtWidgets.QFormLayout()
        # 稳定性 fam 开始测试
        self.btn_stability_fam = QtWidgets.QPushButton("开始测试")
        self.btn_stability_fam.clicked.connect(self.stability_fam)
        self.label_stability_fam = QtWidgets.QLabel()
        self.label_stability_fam.setMinimumHeight(50)
        self.label_stability_fam.setWordWrap(True)

        # 稳定性 rox 开始测试
        self.btn_stability_rox = QtWidgets.QPushButton("开始测试")
        self.btn_stability_rox.clicked.connect(self.stability_rox)
        self.label_stability_rox = QtWidgets.QLabel()
        self.label_stability_rox.setMinimumHeight(50)
        self.label_stability_rox.setWordWrap(True)

        # 稳定性 hex 开始测试
        self.btn_stability_hex = QtWidgets.QPushButton("开始测试")
        self.btn_stability_hex.clicked.connect(self.stability_hex)
        self.label_stability_hex = QtWidgets.QLabel()
        self.label_stability_hex.setMinimumHeight(50)
        self.label_stability_hex.setWordWrap(True)

        # 稳定性 cy5 开始测试
        self.btn_stability_cy5 = QtWidgets.QPushButton("开始测试")
        self.btn_stability_cy5.clicked.connect(self.stability_cy5)
        self.label_stability_cy5 = QtWidgets.QLabel()
        self.label_stability_cy5.setMinimumHeight(50)
        self.label_stability_cy5.setWordWrap(True)


        f_stability.addRow(QtWidgets.QLabel("FAM通道  "), self.btn_stability_fam)
        f_stability.addRow(self.label_stability_fam)
        f_stability.addRow(QtWidgets.QLabel("ROX通道  "), self.btn_stability_rox)
        f_stability.addRow(self.label_stability_rox)
        f_stability.addRow(QtWidgets.QLabel("HEX通道  "), self.btn_stability_hex)
        f_stability.addRow(self.label_stability_hex)
        f_stability.addRow(QtWidgets.QLabel("CY5通道  "), self.btn_stability_cy5)
        f_stability.addRow(self.label_stability_cy5)

        v_stability.addItem(f_stability)
        v_stability.addStretch()

        self.group_stability.setLayout(v_stability)

        # self.module_connect_status_signal.connect(self.)


        # ------------截止率布局--------------
        v_relativity = QtWidgets.QVBoxLayout()
        f_relativity = QtWidgets.QFormLayout()

        # 截止率 fam 开始测试
        self.btn_relativity_fam = QtWidgets.QPushButton("开始测试")
        self.btn_relativity_fam.clicked.connect(self.relativity_fam)
        self.label_relativity_fam = QtWidgets.QLabel()
        self.label_relativity_fam.setMinimumHeight(50)
        self.label_relativity_fam.setWordWrap(True)

        # 截止率 rox 开始测试
        self.btn_relativity_rox = QtWidgets.QPushButton("开始测试")
        self.btn_relativity_rox.clicked.connect(self.relativity_rox)
        self.label_relativity_rox = QtWidgets.QLabel()
        self.label_relativity_rox.setMinimumHeight(50)
        self.label_relativity_rox.setWordWrap(True)

        # 截止率 hex 开始测试
        self.btn_relativity_hex = QtWidgets.QPushButton("开始测试")
        self.btn_relativity_hex.clicked.connect(self.relativity_hex)
        self.label_relativity_hex = QtWidgets.QLabel()
        self.label_relativity_hex.setMinimumHeight(50)
        self.label_relativity_hex.setWordWrap(True)

        # 截止率 cy5 开始测试
        self.btn_relativity_cy5 = QtWidgets.QPushButton("开始测试")
        self.btn_relativity_cy5.clicked.connect(self.relativity_cy5)
        self.label_relativity_cy5 = QtWidgets.QLabel()
        self.label_relativity_cy5.setMinimumHeight(50)
        self.label_relativity_cy5.setWordWrap(True)

        f_relativity.addRow(QtWidgets.QLabel("FAM通道  "), self.btn_relativity_fam)
        f_relativity.addRow(self.label_relativity_fam)
        f_relativity.addRow(QtWidgets.QLabel("ROX通道  "), self.btn_relativity_rox)
        f_relativity.addRow(self.label_relativity_rox)
        f_relativity.addRow(QtWidgets.QLabel("HEX通道  "), self.btn_relativity_hex)
        f_relativity.addRow(self.label_relativity_hex)
        f_relativity.addRow(QtWidgets.QLabel("CY5通道  "), self.btn_relativity_cy5)
        f_relativity.addRow(self.label_relativity_cy5)

        v_relativity.addItem(f_relativity)
        v_relativity.addStretch()

        self.group_relativity.setLayout(v_relativity)



        # ------------灵敏度布局--------------
        v_sensitivity = QtWidgets.QVBoxLayout()
        f_sensitivity= QtWidgets.QFormLayout()
        # 灵敏度 fam 开始测试
        self.btn_sensitivity_fam = QtWidgets.QPushButton("开始测试")
        self.btn_sensitivity_fam.clicked.connect(self.sensitivity_fam)
        self.label_sensitivity_fam = QtWidgets.QLabel()
        self.label_sensitivity_fam.setMinimumHeight(50)
        self.label_sensitivity_fam.setWordWrap(True)

        # 灵敏度 rox 开始测试
        self.btn_sensitivity_rox = QtWidgets.QPushButton("开始测试")
        self.btn_sensitivity_rox.clicked.connect(self.sensitivity_rox)
        self.label_sensitivity_rox = QtWidgets.QLabel()
        self.label_sensitivity_rox.setMinimumHeight(50)
        self.label_sensitivity_rox.setWordWrap(True)

        # 灵敏度 hex 开始测试
        self.btn_sensitivity_hex = QtWidgets.QPushButton("开始测试")
        self.btn_sensitivity_hex.clicked.connect(self.sensitivity_hex)
        self.label_sensitivity_hex = QtWidgets.QLabel()
        self.label_sensitivity_hex.setMinimumHeight(50)
        self.label_sensitivity_hex.setWordWrap(True)

        # 灵敏度 cy5 开始测试
        self.btn_sensitivity_cy5 = QtWidgets.QPushButton("开始测试")
        self.btn_sensitivity_cy5.clicked.connect(self.sensitivity_cy5)
        self.label_sensitivity_cy5 = QtWidgets.QLabel()
        self.label_sensitivity_cy5.setMinimumHeight(50)
        self.label_sensitivity_cy5.setWordWrap(True)

        f_sensitivity.addRow(QtWidgets.QLabel("FAM通道  "), self.btn_sensitivity_fam)
        f_sensitivity.addRow(self.label_sensitivity_fam)
        f_sensitivity.addRow(QtWidgets.QLabel("ROX通道  "), self.btn_sensitivity_rox)
        f_sensitivity.addRow(self.label_sensitivity_rox)
        f_sensitivity.addRow(QtWidgets.QLabel("HEX通道  "), self.btn_sensitivity_hex)
        f_sensitivity.addRow(self.label_sensitivity_hex)
        f_sensitivity.addRow(QtWidgets.QLabel("CY5通道  "), self.btn_sensitivity_cy5)
        f_sensitivity.addRow(self.label_sensitivity_cy5)

        v_sensitivity.addItem(f_sensitivity)
        v_sensitivity.addStretch()

        self.group_sensitivity.setLayout(v_sensitivity)



    def connect_status(self, result):
        if not result:
            self.label_connect_condition.setText(f"<div>模 块 状 态：{self.module_sn} 未 连 接 <b><font size=5 color=red>o</font></b></div>")
        else:
            self.label_connect_condition.setText(f"<div>模 块 状 态：{self.module_sn} 已 连 接 <b><font size=5 color=lime>o</font></b></div>")

    def try_connect_module_com(self):
        self.module_sn = self.sn_module_edit.text()
        if self.module_sn == "":
            Tip("请填写光学模块的SN号", parent=self.window(), width=300)
            return
        connected_result = self.scandedtest.try_connect_optics_module()
        self.connect_signal.emit(connected_result)

    def operation_control(self, result, info):
        if result:
            Alert("处理成功", parent=self.window())
        else:
            Tip(info, parent=self.window(), width=300)

    def stability_fam(self):
        loading = Loading(self.window(), tip="FAM稳定性测试中...")
        thread = threading.Thread(target=self._stability_fam, args=(loading,))
        thread.daemon = True
        thread.start()

    def _stability_fam(self, loading):
        result, info = self.stability_fam_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def stability_fam_ui(self):
        stability_fam_list = self.scandedtest.stabilityfam()
        if not stability_fam_list:
            return False, "fam通道稳定性没有从光学模块中读取到数据"
        stability_fam_avg, stability_fam_cv = self.avgcv(stability_fam_list)
        if not stability_fam_avg:
            return False, "fam通道稳定性光学模块列表数据异常，未能获取到平均值"
        self.label_stability_fam.setText(f"CV 值：  {stability_fam_cv}   结 果：  {self.pf(stability_fam_cv < 0.03)}")
        datapoints_report_dict['stability']['fam'] = [stability_fam_avg, stability_fam_cv, stability_fam_list]
        return True, datapoints_report_dict

    def stability_rox(self):
        loading = Loading(self.window(), tip="ROX稳定性测试中...")
        thread = threading.Thread(target=self._stability_rox, args=(loading,))
        thread.daemon = True
        thread.start()

    def _stability_rox(self, loading):
        result, info = self.stability_rox_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def stability_rox_ui(self):
        stability_rox_list = self.scandedtest.stabilityrox()
        if not stability_rox_list:
            return False, "rox通道稳定性没有从光学模块中读取到数据"
        stability_rox_avg, stability_rox_cv = self.avgcv(stability_rox_list)
        if not stability_rox_avg:
            return False, "rox通道稳定性光学模块列表数据异常，未能获取到平均值"
        self.label_stability_rox.setText(f"CV 值：  {stability_rox_cv}   结 果：  {self.pf(stability_rox_cv < 0.03)}")
        datapoints_report_dict['stability']['rox'] = [stability_rox_avg, stability_rox_cv, stability_rox_list]
        return True, datapoints_report_dict

    def stability_hex(self):
        loading = Loading(self.window(), tip="HEX稳定性测试中...")
        thread = threading.Thread(target=self._stability_hex, args=(loading,))
        thread.daemon = True
        thread.start()

    def _stability_hex(self, loading):
        result, info = self.stability_hex_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def stability_hex_ui(self):
        stability_hex_list = self.scandedtest.stabilityhex()
        if not stability_hex_list:
            return False, "hex通道稳定性没有从光学模块中读取到数据"
        stability_hex_avg, stability_hex_cv = self.avgcv(stability_hex_list)
        if not stability_hex_avg:
            return False, "hex通道稳定性光学模块列表数据异常，未能获取到平均值"
        self.label_stability_hex.setText(f"CV 值：  {stability_hex_cv}   结 果：  {self.pf(stability_hex_cv < 0.03)}")
        datapoints_report_dict['stability']['hex'] = [stability_hex_avg, stability_hex_cv, stability_hex_list]
        return True, datapoints_report_dict

    def stability_cy5(self):
        loading = Loading(self.window(), tip="CY5稳定性测试中...")
        thread = threading.Thread(target=self._stability_cy5, args=(loading,))
        thread.daemon = True
        thread.start()

    def _stability_cy5(self, loading):
        result, info = self.stability_cy5_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def stability_cy5_ui(self):
        stability_cy5_list = self.scandedtest.stabilitycy5()
        if not stability_cy5_list:
            return False, "cy5通道稳定性没有从光学模块中读取到数据"
        stability_cy5_avg, stability_cy5_cv = self.avgcv(stability_cy5_list)
        if not stability_cy5_avg:
            return False, "cy5通道稳定性光学模块列表数据异常，未能获取到平均值"
        self.label_stability_cy5.setText(f"CV 值：  {stability_cy5_cv}   结 果：  {self.pf(stability_cy5_cv < 0.03)}")
        datapoints_report_dict['stability']['cy5'] = [stability_cy5_avg, stability_cy5_cv, stability_cy5_list]
        return True, datapoints_report_dict

    def relativity_fam(self):
        loading = Loading(self.window(), tip="FAM截止率测试中...")
        thread = threading.Thread(target=self._relativity_fam, args=(loading,))
        thread.daemon = True
        thread.start()

    def _relativity_fam(self, loading):
        result, info = self.relativity_fam_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def relativity_fam_ui(self):
        relativity_fam_list = self.scandedtest.relativityfam()
        if not relativity_fam_list:
            return False, "fam通道截止率没有从光学模块中读取到数据"
        relativity_fam_avg, _ = self.avgcv(relativity_fam_list)
        if not relativity_fam_avg:
            return False, "fam通道截止率光学模块列表数据异常，未能获取到平均值"

        if datapoints_report_dict['stability']['fam'][0] == -100:
            self.label_relativity_fam.setText(f"平 均 值： "
                                              f" {relativity_fam_avg}   结 果： "
                                              f" 无 ")
        else:
            stability_minus_relativity_fam = abs(
                round(datapoints_report_dict['stability']['fam'][0] - relativity_fam_avg, 1))
            self.label_relativity_fam.setText(f"差 值：  "
                                              f"{stability_minus_relativity_fam}   结 果：  "
                                              f"{self.pf(stability_minus_relativity_fam < 10)}")
        datapoints_report_dict['relativity']['fam'] = [relativity_fam_avg, relativity_fam_list]
        return True, datapoints_report_dict

    def relativity_rox(self):
        loading = Loading(self.window(), tip="ROX截止率测试中...")
        thread = threading.Thread(target=self._relativity_rox, args=(loading,))
        thread.daemon = True
        thread.start()

    def _relativity_rox(self, loading):
        result, info = self.relativity_rox_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def relativity_rox_ui(self):
        relativity_rox_list = self.scandedtest.relativityrox()
        if not relativity_rox_list:
            return False, "rox通道截止率没有从光学模块中读取到数据"
        relativity_rox_avg, _ = self.avgcv(relativity_rox_list)
        if not relativity_rox_avg:
            return False, "rox通道截止率光学模块列表数据异常，未能获取到平均值"

        if datapoints_report_dict['stability']['rox'][0] == -100:
            self.label_relativity_rox.setText(f"平 均 值：  "
                                              f" {relativity_rox_avg}   结 果：  "
                                              f" 无 ")
        else:
            stability_minus_relativity_rox = abs(
                round(datapoints_report_dict['stability']['rox'][0] - relativity_rox_avg, 1))
            self.label_relativity_rox.setText(f"差 值：  "
                                              f"{stability_minus_relativity_rox}   结 果：  "
                                              f"{self.pf(stability_minus_relativity_rox < 10)}")
        datapoints_report_dict['relativity']['rox'] = [relativity_rox_avg, relativity_rox_list]
        return True, datapoints_report_dict

    def relativity_hex(self):
        loading = Loading(self.window(), tip="HEX截止率测试中...")
        thread = threading.Thread(target=self._relativity_hex, args=(loading,))
        thread.daemon = True
        thread.start()

    def _relativity_hex(self, loading):
        result, info = self.relativity_hex_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def relativity_hex_ui(self):
        relativity_hex_list = self.scandedtest.relativityhex()
        if not relativity_hex_list:
            return False, "hex通道截止率没有从光学模块中读取到数据"
        relativity_hex_avg, _ = self.avgcv(relativity_hex_list)
        if not relativity_hex_avg:
            return False, "hex通道截止率光学模块列表数据异常，未能获取到平均值"

        if datapoints_report_dict['stability']['hex'][0] == -100:
            self.label_relativity_hex.setText(f"平 均 值：  "
                                              f" {relativity_hex_avg}   结 果：  "
                                              f" 无 ")
        else:
            stability_minus_relativity_hex = abs(
                round(datapoints_report_dict['stability']['hex'][0] - relativity_hex_avg, 1))
            self.label_relativity_hex.setText(f"差 值：  "
                                              f"{stability_minus_relativity_hex}   结 果：  "
                                              f"{self.pf(stability_minus_relativity_hex < 10)}")
        datapoints_report_dict['relativity']['hex'] = [relativity_hex_avg, relativity_hex_list]
        return True, datapoints_report_dict

    def relativity_cy5(self):
        loading = Loading(self.window(), tip="CY5截止率测试中...")
        thread = threading.Thread(target=self._relativity_cy5, args=(loading,))
        thread.daemon = True
        thread.start()

    def _relativity_cy5(self, loading):
        result, info = self.relativity_cy5_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def relativity_cy5_ui(self):
        relativity_cy5_list = self.scandedtest.relativitycy5()
        if not relativity_cy5_list:
            return False, "cy5通道截止率没有从光学模块中读取到数据"
        relativity_cy5_avg, _ = self.avgcv(relativity_cy5_list)
        if not relativity_cy5_avg:
            return False, "cy5通道截止率光学模块列表数据异常，未能获取到平均值"

        if datapoints_report_dict['stability']['cy5'][0] == -100:
            self.label_relativity_cy5.setText(f"平 均 值：  "
                                              f" {relativity_cy5_avg}   结 果：  "
                                              f" 无 ")
        else:
            stability_minus_relativity_cy5 = abs(
                round(datapoints_report_dict['stability']['cy5'][0] - relativity_cy5_avg, 1))
            self.label_relativity_cy5.setText(f"差 值：  "
                                              f"{stability_minus_relativity_cy5}   结 果：  "
                                              f"{self.pf(stability_minus_relativity_cy5 < 10)}")
        datapoints_report_dict['relativity']['cy5'] = [relativity_cy5_avg, relativity_cy5_list]
        return True, datapoints_report_dict

    def sensitivity_fam(self):
        loading = Loading(self.window(), tip="FAM灵敏度测试中...")
        thread = threading.Thread(target=self._sensitivity_fam, args=(loading, ))
        thread.daemon = True
        thread.start()

    def _sensitivity_fam(self, loading):
        result, info = self.sensitivity_fam_ui()
        self.operation_signal.emit(result, info)

        loading.close_loading()

    def sensitivity_fam_ui(self):
        sensitivity_fam_list = self.scandedtest.sensitivityfam()
        if not sensitivity_fam_list:
            return False, f"fam通道灵敏度没有从光学模块中读取到数据"
        sensitivity_fam_avg, _ = self.avgcv(sensitivity_fam_list)
        if not sensitivity_fam_avg:
            return False, f"fam通道灵敏度光学模块列表数据异常，未能获取到平均值"
        self.label_sensitivity_fam.setText(f"荧 光 值："
                                           f"{sensitivity_fam_avg}   结 果："
                                           f"{self.pf(900 < sensitivity_fam_avg < 2450)}")
        datapoints_report_dict['sensitivity']['fam'] = [sensitivity_fam_avg, sensitivity_fam_list]

        return True, datapoints_report_dict

    def sensitivity_rox(self):
        loading = Loading(self.window(), tip="ROX灵敏度测试中...")
        thread = threading.Thread(target=self._sensitivity_rox, args=(loading,))
        thread.daemon = True
        thread.start()

    def _sensitivity_rox(self, loading):
        result, info = self.sensitivity_rox_ui()

        self.operation_signal.emit(result, info)
        loading.close_loading()

    def sensitivity_rox_ui(self):
        sensitivity_rox_list = self.scandedtest.sensitivityrox()
        if not sensitivity_rox_list:
            return False, "rox通道灵敏度没有从光学模块中读取到数据"
        sensitivity_rox_avg, _ = self.avgcv(sensitivity_rox_list)
        if not sensitivity_rox_avg:
            return False, "rox通道灵敏度光学模块列表数据异常，未能获取到平均值"
        self.label_sensitivity_rox.setText(f"荧 光 值："
                                           f"{sensitivity_rox_avg}  结 果："
                                           f"{self.pf(1000 < sensitivity_rox_avg < 2450)}")
        datapoints_report_dict['sensitivity']['rox'] = [sensitivity_rox_avg, sensitivity_rox_list]
        return True, datapoints_report_dict

    def sensitivity_hex(self):
        loading = Loading(self.window(), tip="HEX灵敏度测试中...")
        thread = threading.Thread(target=self._sensitivity_hex, args=(loading,))
        thread.daemon = True
        thread.start()

    def _sensitivity_hex(self, loading):
        result, info = self.sensitivity_hex_ui()

        self.operation_signal.emit(result, info)
        loading.close_loading()

    def sensitivity_hex_ui(self):
        sensitivity_hex_list = self.scandedtest.sensitivityhex()
        if not sensitivity_hex_list:
            return False, "hex通道灵敏度没有从光学模块中读取到数据"
        sensitivity_hex_avg, _ = self.avgcv(sensitivity_hex_list)
        if not sensitivity_hex_avg:
            return False, "hex通道灵敏度光学模块列表数据异常，未能获取到平均值"

        self.label_sensitivity_hex.setText(f"荧 光 值："
                                           f"{sensitivity_hex_avg}  结 果："
                                           f"{self.pf(1000 < sensitivity_hex_avg < 2450)}")
        datapoints_report_dict['sensitivity']['hex'] = [sensitivity_hex_avg, sensitivity_hex_list]
        return True, datapoints_report_dict

    def sensitivity_cy5(self):
        loading = Loading(self.window(), tip="CY5灵敏度测试中...")
        thread = threading.Thread(target=self._sensitivity_cy5, args=(loading,))
        thread.daemon = True
        thread.start()

    def _sensitivity_cy5(self, loading):
        result, info = self.sensitivity_cy5_ui()

        self.operation_signal.emit(result, info)
        loading.close_loading()

    def sensitivity_cy5_ui(self):
        sensitivity_cy5_list = self.scandedtest.sensitivitycy5()
        if not sensitivity_cy5_list:
            return False, "cy5通道灵敏度没有从光学模块中读取到数据"
        sensitivity_cy5_avg, _ = self.avgcv(sensitivity_cy5_list)
        if not sensitivity_cy5_avg:
            return False, "cy5通道灵敏度光学模块列表数据异常，未能获取到平均值"

        self.label_sensitivity_cy5.setText(f"荧 光 值："
                                           f"{sensitivity_cy5_avg}  结 果："
                                           f"{self.pf(1000 < sensitivity_cy5_avg < 2450)}")
        datapoints_report_dict['sensitivity']['cy5'] = [sensitivity_cy5_avg, sensitivity_cy5_list]
        return True, datapoints_report_dict

    def export_report(self):
        """
        生成报告并导出
        """
        try:
            wb_datapoint = load_workbook(datapoint_template)
            ws_deal_result = wb_datapoint['处理结果']
            ws_original_data = wb_datapoint['原始数据']
            # _________________稳定性____________________________
            ws_deal_result.cell(5, 1, self.module_sn)
            ws_deal_result.cell(10, 1, self.module_sn)
            ws_deal_result.cell(15, 1, self.module_sn)
            ws_deal_result.cell(5, 2, datapoints_report_dict['stability']['fam'][0])
            ws_deal_result.cell(10, 2, datapoints_report_dict['stability']['fam'][0])
            ws_deal_result.cell(15, 2, datapoints_report_dict['stability']['fam'][0])
            ws_deal_result.cell(5, 3, datapoints_report_dict['stability']['fam'][1])
            ws_deal_result.cell(5, 4, datapoints_report_dict['stability']['rox'][0])
            ws_deal_result.cell(10, 5, datapoints_report_dict['stability']['rox'][0])
            ws_deal_result.cell(15, 5, datapoints_report_dict['stability']['rox'][0])
            ws_deal_result.cell(5, 5, datapoints_report_dict['stability']['rox'][1])
            ws_deal_result.cell(5, 6, datapoints_report_dict['stability']['hex'][0])
            ws_deal_result.cell(10, 8, datapoints_report_dict['stability']['hex'][0])
            ws_deal_result.cell(15, 8, datapoints_report_dict['stability']['hex'][0])
            ws_deal_result.cell(5, 7, datapoints_report_dict['stability']['hex'][1])
            ws_deal_result.cell(5, 8, datapoints_report_dict['stability']['cy5'][0])
            ws_deal_result.cell(10, 11, datapoints_report_dict['stability']['cy5'][0])
            ws_deal_result.cell(15, 11, datapoints_report_dict['stability']['cy5'][0])
            ws_deal_result.cell(5, 9, datapoints_report_dict['stability']['cy5'][1])
            for index, datapoint in enumerate(datapoints_report_dict['stability']['fam'][2]):
                ws_original_data.cell(index + 2, 2, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['stability']['rox'][2]):
                ws_original_data.cell(index + 2, 3, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['stability']['hex'][2]):
                ws_original_data.cell(index + 2, 4, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['stability']['cy5'][2]):
                ws_original_data.cell(index + 2, 5, datapoint)

            # _________________截止率____________________________
            ws_deal_result.cell(10, 3, datapoints_report_dict['relativity']['fam'][0])
            ws_deal_result.cell(10, 6, datapoints_report_dict['relativity']['rox'][0])
            ws_deal_result.cell(10, 9, datapoints_report_dict['relativity']['hex'][0])
            ws_deal_result.cell(10, 12, datapoints_report_dict['relativity']['cy5'][0])
            for index, datapoint in enumerate(datapoints_report_dict['relativity']['fam'][1]):
                ws_original_data.cell(index + 2, 6, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['relativity']['rox'][1]):
                ws_original_data.cell(index + 2, 7, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['relativity']['hex'][1]):
                ws_original_data.cell(index + 2, 8, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['relativity']['cy5'][1]):
                ws_original_data.cell(index + 2, 9, datapoint)

            # _________________灵敏度____________________________
            ws_deal_result.cell(15, 3, datapoints_report_dict['sensitivity']['fam'][0])
            ws_deal_result.cell(15, 6, datapoints_report_dict['sensitivity']['rox'][0])
            ws_deal_result.cell(15, 9, datapoints_report_dict['sensitivity']['hex'][0])
            ws_deal_result.cell(15, 12, datapoints_report_dict['sensitivity']['cy5'][0])
            for index, datapoint in enumerate(datapoints_report_dict['sensitivity']['fam'][1]):
                ws_original_data.cell(index + 2, 10, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['sensitivity']['rox'][1]):
                ws_original_data.cell(index + 2, 11, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['sensitivity']['hex'][1]):
                ws_original_data.cell(index + 2, 12, datapoint)
            for index, datapoint in enumerate(datapoints_report_dict['sensitivity']['cy5'][1]):
                ws_original_data.cell(index + 2, 13, datapoint)

            export_datapoints_path = get_file_absolute_path(f'/report/{self.module_sn}光学模块报告.xlsx')
            wb_datapoint.save(export_datapoints_path)
            Alert("导出成功", parent=self.window())
            return True, "报告生成成功"
        except Exception as err:
            Tip(f"报告生成失败{err}", parent=self.window(), width=300)
            return False, f"{err}报告生成失败"

    def avgcv(self, conversion_formula_mv_list):
        """
        返回每次测试的值得平均值和CV值
        :return:
        """
        avg = round(np.mean(conversion_formula_mv_list), 1)
        cv = round(np.std(conversion_formula_mv_list) / np.mean(conversion_formula_mv_list), 3)
        return  avg, cv

    def pf(self, s: bool):
        return "合 格" if s else "不 合 格"

if __name__ == '__main__':
    multiprocessing.freeze_support()
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
